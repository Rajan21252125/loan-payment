from flask import Flask, render_template, request, redirect, flash
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "secret123"

FILE = "data.xlsx"
LOAN_AMOUNT = 3500000
EMI = 29999


def safe_float(val):
    try:
        return float(val)
    except:
        return 0


# Create Excel file if not exists
if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"
    ws.append([
        "Month", "EMI", "Prepayment",
        "Flexi", "Large", "Mid", "Small",
        "Note"
    ])
    wb.save(FILE)


@app.route("/", methods=["GET", "POST"])
def index():
    wb = load_workbook(FILE)
    ws = wb.active

    data = list(ws.iter_rows(min_row=2, values_only=True))

    # ADD ENTRY
    if request.method == "POST":
        month = request.form["month"]

        # Prevent duplicate month
        if month in [r[0] for r in data]:
            flash("⚠️ Entry for this month already exists!")
            return redirect("/")

        ws.append([
            month,
            request.form["emi"],
            safe_float(request.form.get("prepayment")),
            safe_float(request.form.get("flexi")),
            safe_float(request.form.get("large")),
            safe_float(request.form.get("mid")),
            safe_float(request.form.get("small")),
            request.form.get("note")
        ])

        wb.save(FILE)
        return redirect("/")

    # CALCULATIONS
    total_emi = sum(EMI for r in data if r[1] == "Y")
    total_prepayment = sum(safe_float(r[2]) for r in data)

    total_flexi = sum(safe_float(r[3]) for r in data)
    total_large = sum(safe_float(r[4]) for r in data)
    total_mid = sum(safe_float(r[5]) for r in data)
    total_small = sum(safe_float(r[6]) for r in data)

    # Loan Remaining
    paid_amount = total_emi + total_prepayment
    remaining_loan = max(LOAN_AMOUNT - paid_amount, 0)

    months_left = int(remaining_loan / EMI) if EMI else 0

    # Month dropdown
    months = []
    for y in range(2026, 2039):
        for m in ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]:
            months.append(f"{m} {y}")

    current_month = datetime.now().strftime("%b %Y")

    return render_template(
        "index.html",
        data=data,
        months=months,
        current_month=current_month,
        total_emi=total_emi,
        total_prepayment=total_prepayment,
        total_flexi=total_flexi,
        total_large=total_large,
        total_mid=total_mid,
        total_small=total_small,
        remaining_loan=remaining_loan,
        months_left=months_left
    )


# DELETE
@app.route("/delete/<int:index>")
def delete(index):
    wb = load_workbook(FILE)
    ws = wb.active
    ws.delete_rows(index + 2)
    wb.save(FILE)
    return redirect("/")


# EDIT
@app.route("/edit/<int:index>", methods=["POST"])
def edit(index):
    wb = load_workbook(FILE)
    ws = wb.active

    row = index + 2

    ws.cell(row=row, column=2).value = request.form["emi"]
    ws.cell(row=row, column=3).value = safe_float(request.form["prepayment"])
    ws.cell(row=row, column=4).value = safe_float(request.form["flexi"])
    ws.cell(row=row, column=5).value = safe_float(request.form["large"])
    ws.cell(row=row, column=6).value = safe_float(request.form["mid"])
    ws.cell(row=row, column=7).value = safe_float(request.form["small"])
    ws.cell(row=row, column=8).value = request.form["note"]

    wb.save(FILE)
    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)