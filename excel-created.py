from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

# =========================
# 1. MONTHLY TRACKER
# =========================
ws = wb.active
ws.title = "Monthly Tracker"

headers = [
    "Year", "Month",
    "Planned EMI", "EMI Paid? (Y/N)", "Actual EMI",
    "Extra EMI Paid? (Y/N)", "Extra EMI Amount",
    "SIP Done? (Y/N)", "SIP Amount",
    "Prepayment",
]

ws.append(headers)

for col in range(1, len(headers)+1):
    ws.cell(row=1, column=col).font = Font(bold=True)

months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

row = 2
for y in range(2026, 2038):
    for m in months:
        ws.append([
            y, m,
            29999, "", "",   # EMI
            "", "",          # Extra EMI
            "", 5000,        # SIP
            ""               # Prepayment
        ])
        row += 1

# =========================
# 2. LOAN TRACKER (FIXED)
# =========================
loan_ws = wb.create_sheet("Loan Tracker")

loan_ws.append([
    "Month", "Opening Balance", "EMI Used",
    "Extra EMI", "Prepayment",
    "Interest", "Principal", "Closing Balance"
])

loan_amount = 3500000
rate = 8.4 / 100 / 12

loan_ws.cell(row=2, column=2).value = loan_amount

for i in range(2, 200):

    # EMI only if paid
    loan_ws.cell(row=i, column=3).value = f"""=IF('Monthly Tracker'!D{i}="Y",'Monthly Tracker'!E{i},0)"""

    # Extra EMI only if paid
    loan_ws.cell(row=i, column=4).value = f"""=IF('Monthly Tracker'!F{i}="Y",'Monthly Tracker'!G{i},0)"""

    # Prepayment always direct
    loan_ws.cell(row=i, column=5).value = f"='Monthly Tracker'!J{i}"

    # Interest
    loan_ws.cell(row=i, column=6).value = f"=B{i}*{rate}"

    # Principal calculation
    loan_ws.cell(row=i, column=7).value = f"=C{i}+D{i}+E{i}-F{i}"

    # Closing balance
    loan_ws.cell(row=i, column=8).value = f"=B{i}-G{i}"

    if i < 199:
        loan_ws.cell(row=i+1, column=2).value = f"=H{i}"

# =========================
# 3. SIP TRACKER (FIXED)
# =========================
sip_ws = wb.create_sheet("SIP Growth")

sip_ws.append(["Month", "SIP Used", "Total Invested", "Value"])

rate = 0.12 / 12

for i in range(1, 200):

    # SIP only if done
    sip_ws.cell(row=i+1, column=2).value = f"""=IF('Monthly Tracker'!H{i+1}="Y",'Monthly Tracker'!I{i+1},0)"""

    # Total invested
    sip_ws.cell(row=i+1, column=3).value = f"=SUM(B2:B{i+1})"

    # Growth
    sip_ws.cell(row=i+1, column=4).value = f"=B{i+1}*(1+{rate})^{i}"

# SAVE
wb.save("realistic_financial_tracker.xlsx")

print("Updated file created!")